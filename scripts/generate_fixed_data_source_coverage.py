from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
MANIFEST_PATH = ROOT / "data" / "catalogue" / "fixed_data_source_manifest.json"
SOURCE_ROOT = ROOT / "data" / "catalogue" / "source_library"
REPORT_JSON = ROOT / "data" / "catalogue" / "fixed_data_source_coverage.json"
REPORT_XLSX = ROOT / "Fixed_Data_Source_Coverage.xlsx"


def autosize_columns(worksheet) -> None:
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 52)


def style_sheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_columns(worksheet)


def main() -> None:
    manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))

    sources_by_system = {}
    for file in SOURCE_ROOT.glob("*\\_source_links.json"):
        try:
            data = json.loads(file.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(data, list):
            continue
        for entry in data:
            system_id = entry.get("system_id")
            if system_id:
                sources_by_system[system_id] = entry.get("sources", [])

    detailed_rows = []
    supplier_rollup = Counter()
    sourced_supplier_rollup = Counter()
    missing_rows = []

    for item in manifest:
        supplier_name = item.get("supplier_name") or "Unassigned"
        system_id = item.get("system_id", "")
        sources = sources_by_system.get(system_id, [])
        source_count = len(sources)
        has_sources = source_count > 0
        supplier_rollup[supplier_name] += 1
        if has_sources:
            sourced_supplier_rollup[supplier_name] += 1
        else:
            missing_rows.append(
                {
                    "Supplier Name": supplier_name,
                    "System ID": system_id,
                    "System Name": item.get("system_name", ""),
                    "System Category": item.get("system_category", ""),
                    "Source Folder": item.get("source_folder", ""),
                    "Missing Source Types": "operative technique / IFU / catalogue not yet logged",
                    "Notes": item.get("notes", ""),
                }
            )
        detailed_rows.append(
            {
                "Supplier Name": supplier_name,
                "System ID": system_id,
                "System Name": item.get("system_name", ""),
                "System Category": item.get("system_category", ""),
                "Mapping Count": item.get("mapping_count", 0),
                "Source Count": source_count,
                "Source Status": "sourced" if has_sources else "not sourced",
                "Source Folder": item.get("source_folder", ""),
            }
        )

    supplier_rows = []
    for supplier_name, total in sorted(supplier_rollup.items(), key=lambda pair: (-pair[1], pair[0])):
        sourced = sourced_supplier_rollup.get(supplier_name, 0)
        supplier_rows.append(
            {
                "Supplier Name": supplier_name,
                "Systems Total": total,
                "Systems With Sources": sourced,
                "Systems Missing Sources": total - sourced,
                "Coverage %": round((sourced / total) * 100, 1) if total else 0,
            }
        )

    report = {
        "summary": {
            "systems_total": len(manifest),
            "systems_with_sources": sum(1 for row in detailed_rows if row["Source Status"] == "sourced"),
            "systems_missing_sources": sum(1 for row in detailed_rows if row["Source Status"] == "not sourced"),
        },
        "supplier_coverage": supplier_rows,
        "system_coverage": detailed_rows,
        "missing_sources": missing_rows,
    }
    REPORT_JSON.write_text(json.dumps(report, indent=2, ensure_ascii=True), encoding="utf-8")

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "SUMMARY"
    ws_summary.append(["Metric", "Value"])
    for key, value in report["summary"].items():
        ws_summary.append([key, value])
    style_sheet(ws_summary)

    ws_suppliers = wb.create_sheet("SUPPLIER_COVERAGE")
    ws_suppliers.append(list(supplier_rows[0].keys()) if supplier_rows else ["Supplier Name"])
    for row in supplier_rows:
        ws_suppliers.append(list(row.values()))
    style_sheet(ws_suppliers)

    ws_systems = wb.create_sheet("SYSTEM_COVERAGE")
    ws_systems.append(list(detailed_rows[0].keys()) if detailed_rows else ["System ID"])
    for row in detailed_rows:
        ws_systems.append(list(row.values()))
    style_sheet(ws_systems)

    ws_missing = wb.create_sheet("MISSING_SOURCES")
    ws_missing.append(list(missing_rows[0].keys()) if missing_rows else ["System ID"])
    for row in missing_rows:
        ws_missing.append(list(row.values()))
    style_sheet(ws_missing)

    wb.save(REPORT_XLSX)
    print(REPORT_JSON)
    print(REPORT_XLSX)


if __name__ == "__main__":
    main()
