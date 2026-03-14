from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MASTER_WORKBOOK = ROOT / "PrepSight_Fixed_Data_Inventory.xlsx"
TARGET_SHEETS = [
    "PROCEDURES",
    "SUPPLIERS",
    "SYSTEMS",
    "TRAYS",
    "COMPONENTS",
    "PRODUCTS",
    "SYSTEM_MAPPINGS",
]


def normalize_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_sheet(workbook_path: Path, sheet_name: str) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return {"headers": [], "rows": [], "count": 0}
    headers = [normalize_value(value) for value in rows[0]]
    data_rows = []
    for row in rows[1:]:
        record = {}
        for index, header in enumerate(headers):
            record[header] = normalize_value(row[index] if index < len(row) else "")
        if any(record.values()):
            data_rows.append(record)
    return {"headers": headers, "rows": data_rows, "count": len(data_rows)}


def main() -> None:
    parser = argparse.ArgumentParser(description="Export fixed-data master workbook to JSON.")
    parser.add_argument("--master", type=Path, default=DEFAULT_MASTER_WORKBOOK, help="Path to master workbook.")
    args = parser.parse_args()

    dataset = {sheet: read_sheet(args.master, sheet) for sheet in TARGET_SHEETS}
    print(json.dumps(dataset, ensure_ascii=True))


if __name__ == "__main__":
    main()
