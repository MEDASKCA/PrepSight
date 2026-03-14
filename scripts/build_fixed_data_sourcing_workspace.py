from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SYSTEMS_PATH = ROOT / "data" / "systems" / "systems.json"
SUPPLIERS_PATH = ROOT / "data" / "suppliers" / "suppliers.json"
MAPPINGS_PATH = ROOT / "data" / "systems" / "procedure_variant_system_map.json"

SOURCE_ROOT = ROOT / "data" / "catalogue" / "source_library"
MANIFEST_JSON = ROOT / "data" / "catalogue" / "fixed_data_source_manifest.json"
WORKLIST_XLSX = ROOT / "Fixed_Data_Sourcing_Worklist.xlsx"


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-") or "unknown"


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def autosize_columns(worksheet) -> None:
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def style_sheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_columns(worksheet)


def ensure_text_file(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        path.write_text(text, encoding="utf-8")


def main() -> None:
    systems = load_json(SYSTEMS_PATH)
    suppliers = load_json(SUPPLIERS_PATH)
    mappings = load_json(MAPPINGS_PATH)

    supplier_name_by_id = {row["id"]: row.get("name", "") for row in suppliers}
    mapping_counts = Counter(row.get("system_id", "") for row in mappings)

    SOURCE_ROOT.mkdir(parents=True, exist_ok=True)

    manifest = []
    systems_rows = []
    supplier_rows = []

    supplier_counts = Counter()
    systems_by_supplier = defaultdict(list)

    for supplier in suppliers:
        supplier_id = supplier.get("id", "")
        supplier_name = supplier.get("name", "")
        supplier_slug = slug(supplier_name)
        supplier_dir = SOURCE_ROOT / supplier_slug
        supplier_dir.mkdir(parents=True, exist_ok=True)
        ensure_text_file(
            supplier_dir / "README.txt",
            f"Supplier: {supplier_name}\nSupplier ID: {supplier_id}\nStore operative techniques, IFUs, and catalogues for this supplier in subfolders by system.\n",
        )

    for system in systems:
        supplier_id = system.get("supplier_id", "")
        supplier_name = supplier_name_by_id.get(supplier_id, "")
        supplier_slug = slug(supplier_name or "unassigned")
        system_id = system.get("id", "")
        system_name = system.get("name", "")
        system_slug = slug(system_name or system_id)
        system_dir = SOURCE_ROOT / supplier_slug / f"{system_slug}__{system_id}"
        system_dir.mkdir(parents=True, exist_ok=True)
        ensure_text_file(
            system_dir / "README.txt",
            (
                f"System: {system_name}\n"
                f"System ID: {system_id}\n"
                f"Supplier: {supplier_name or 'Unassigned'}\n"
                f"Supplier ID: {supplier_id}\n"
                f"Store operative techniques, IFUs, catalogues, and evidence PDFs in this folder.\n"
            ),
        )

        mapping_count = mapping_counts.get(system_id, 0)
        source_entry = {
            "supplier_id": supplier_id,
            "supplier_name": supplier_name,
            "system_id": system_id,
            "system_name": system_name,
            "system_category": system.get("category", ""),
            "system_type": system.get("system_type", ""),
            "mapping_count": mapping_count,
            "source_folder": str(system_dir.relative_to(ROOT)),
            "source_status": "not_started",
            "operative_technique_found": "",
            "ifu_found": "",
            "catalogue_found": "",
            "source_links": [],
            "notes": "",
        }
        manifest.append(source_entry)
        supplier_counts[supplier_id] += 1
        systems_by_supplier[supplier_id].append(source_entry)
        systems_rows.append(
            [
                system_id,
                system_name,
                supplier_id,
                supplier_name,
                system.get("category", ""),
                system.get("system_type", ""),
                mapping_count,
                source_entry["source_folder"],
                source_entry["source_status"],
                "",
            ]
        )

    for supplier in suppliers:
        supplier_id = supplier.get("id", "")
        supplier_name = supplier.get("name", "")
        supplier_rows.append(
            [
                supplier_id,
                supplier_name,
                supplier.get("status", ""),
                supplier_counts.get(supplier_id, 0),
                str((SOURCE_ROOT / slug(supplier_name)).relative_to(ROOT)),
                "",
            ]
        )

    MANIFEST_JSON.write_text(json.dumps(manifest, indent=2, ensure_ascii=True), encoding="utf-8")

    workbook = Workbook()
    ws_suppliers = workbook.active
    ws_suppliers.title = "SUPPLIER_WORKLIST"
    ws_suppliers.append(
        [
            "Supplier ID",
            "Supplier Name",
            "Supplier Status",
            "System Count",
            "Source Folder",
            "Notes",
        ]
    )
    for row in supplier_rows:
        ws_suppliers.append(row)
    style_sheet(ws_suppliers)

    ws_systems = workbook.create_sheet("SYSTEM_WORKLIST")
    ws_systems.append(
        [
            "System ID",
            "System Name",
            "Supplier ID",
            "Supplier Name",
            "System Category",
            "System Type",
            "Mapping Count",
            "Source Folder",
            "Source Status",
            "Notes",
        ]
    )
    for row in systems_rows:
        ws_systems.append(row)
    style_sheet(ws_systems)

    ws_summary = workbook.create_sheet("SUMMARY")
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Suppliers", len(suppliers)])
    ws_summary.append(["Systems", len(systems)])
    ws_summary.append(["System mappings", len(mappings)])
    ws_summary.append(["Source root", str(SOURCE_ROOT.relative_to(ROOT))])
    ws_summary.append(["Manifest JSON", str(MANIFEST_JSON.relative_to(ROOT))])
    style_sheet(ws_summary)

    workbook.save(WORKLIST_XLSX)

    print(str(WORKLIST_XLSX))
    print(str(MANIFEST_JSON))


if __name__ == "__main__":
    main()
