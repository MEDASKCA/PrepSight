from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MASTER_WORKBOOK = ROOT / "PrepSight_Fixed_Data_Inventory.xlsx"
DEFAULT_SCHEMA = ROOT / "data" / "catalogue" / "fixed_data_extraction_schema.json"
TARGET_SHEETS = [
    "PROCEDURES",
    "SUPPLIERS",
    "SYSTEMS",
    "TRAYS",
    "COMPONENTS",
    "PRODUCTS",
    "SYSTEM_MAPPINGS",
]


def normalize_text(value: str) -> str:
    value = re.sub(r"\s+", " ", (value or "").strip())
    return value


def normalize_key(value: str) -> str:
    value = normalize_text(value).lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def slug_token(value: str) -> str:
    token = normalize_key(value).upper()
    return token or "UNKNOWN"


def make_id(prefix: str, seed: str, used_ids: set[str]) -> str:
    base = f"{prefix}_{slug_token(seed)}"
    candidate = base
    suffix = 2
    while candidate in used_ids:
        candidate = f"{base}_{suffix}"
        suffix += 1
    used_ids.add(candidate)
    return candidate


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def validate_schema_shape(payload: dict[str, Any], schema: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    required_top = schema.get("required", [])
    properties = schema.get("properties", {})

    for key in required_top:
        if key not in payload:
            errors.append(f"Missing top-level section: {key}")

    for key, config in properties.items():
        if key not in payload:
            continue
        value = payload[key]
        expected_type = config.get("type")
        if expected_type == "object" and not isinstance(value, dict):
            errors.append(f"Section {key} must be an object.")
            continue
        if expected_type == "array" and not isinstance(value, list):
            errors.append(f"Section {key} must be an array.")
            continue

        if expected_type == "object":
            for field in config.get("required", []):
                if field not in value:
                    errors.append(f"Section {key} is missing field {field}.")
        elif expected_type == "array":
            item_required = config.get("items", {}).get("required", [])
            for index, item in enumerate(value):
                if not isinstance(item, dict):
                    errors.append(f"{key}[{index}] must be an object.")
                    continue
                for field in item_required:
                    if field not in item:
                        errors.append(f"{key}[{index}] is missing field {field}.")
    return errors


def sheet_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    data_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        if row is None:
            continue
        record = {}
        for idx, header in enumerate(headers):
            record[header] = row[idx] if idx < len(row) else None
        if any(value not in ("", None) for value in record.values()):
            data_rows.append(record)
    workbook.close()
    return data_rows


@dataclass
class MasterIndexes:
    procedures: list[dict[str, Any]]
    suppliers: list[dict[str, Any]]
    systems: list[dict[str, Any]]
    trays: list[dict[str, Any]]
    components: list[dict[str, Any]]
    products: list[dict[str, Any]]
    mappings: list[dict[str, Any]]
    supplier_by_name: dict[str, list[dict[str, Any]]]
    system_by_name: dict[str, list[dict[str, Any]]]
    component_by_name: dict[str, list[dict[str, Any]]]
    procedure_by_name_variant: dict[tuple[str, str], list[dict[str, Any]]]
    procedure_by_name_only: dict[str, list[dict[str, Any]]]


def load_master_indexes(workbook_path: Path) -> MasterIndexes:
    procedures = sheet_rows(workbook_path, "PROCEDURES")
    suppliers = sheet_rows(workbook_path, "SUPPLIERS")
    systems = sheet_rows(workbook_path, "SYSTEMS")
    trays = sheet_rows(workbook_path, "TRAYS")
    components = sheet_rows(workbook_path, "COMPONENTS")
    products = sheet_rows(workbook_path, "PRODUCTS")
    mappings = sheet_rows(workbook_path, "SYSTEM_MAPPINGS")

    supplier_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in suppliers:
        supplier_by_name[normalize_key(str(row.get("Supplier Name", "")))].append(row)

    system_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in systems:
        system_by_name[normalize_key(str(row.get("System Name", "")))].append(row)

    component_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in components:
        component_by_name[normalize_key(str(row.get("Component Name", "")))].append(row)

    procedure_by_name_variant: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    procedure_by_name_only: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in procedures:
        name_key = normalize_key(str(row.get("Procedure", "")))
        variant_key = normalize_key(str(row.get("Procedure Variant", "")))
        procedure_by_name_variant[(name_key, variant_key)].append(row)
        procedure_by_name_only[name_key].append(row)

    return MasterIndexes(
        procedures=procedures,
        suppliers=suppliers,
        systems=systems,
        trays=trays,
        components=components,
        products=products,
        mappings=mappings,
        supplier_by_name=supplier_by_name,
        system_by_name=system_by_name,
        component_by_name=component_by_name,
        procedure_by_name_variant=procedure_by_name_variant,
        procedure_by_name_only=procedure_by_name_only,
    )


class Importer:
    def __init__(self, master: MasterIndexes, payload: dict[str, Any]) -> None:
        self.master = master
        self.payload = payload
        self.used_supplier_ids = {str(row.get("Supplier ID", "")) for row in master.suppliers if row.get("Supplier ID")}
        self.used_system_ids = {str(row.get("System ID", "")) for row in master.systems if row.get("System ID")}
        self.used_tray_ids = {str(row.get("Tray ID", "")) for row in master.trays if row.get("Tray ID")}
        self.used_component_ids = {str(row.get("Component ID", "")) for row in master.components if row.get("Component ID")}
        self.used_mapping_ids = {str(row.get("Mapping ID", "")) for row in master.mappings if row.get("Mapping ID")}

        self.generated_suppliers: dict[str, list[str]] = {}
        self.generated_systems: dict[str, list[str]] = {}
        self.generated_components: dict[str, list[str]] = {}

        self.ready: dict[str, list[dict[str, Any]]] = defaultdict(list)
        self.needs_review: dict[str, list[dict[str, Any]]] = defaultdict(list)
        self.rejected: dict[str, list[dict[str, Any]]] = defaultdict(list)
        self.duplicate_warnings: list[dict[str, Any]] = []
        self.missing_field_warnings: list[dict[str, Any]] = []
        self.bad_link_warnings: list[dict[str, Any]] = []
        self.exception_report: list[dict[str, Any]] = []

    def add_issue(self, bucket: dict[str, list[dict[str, Any]]], entity: str, row: dict[str, Any], reason: str) -> None:
        bucket[entity].append({"row": row, "reason": reason})
        self.exception_report.append({"entity": entity, "reason": reason, "row": row})

    def find_existing_system_id(self, system_name: str, supplier_name: str) -> tuple[str | None, str | None]:
        system_key = normalize_key(system_name)
        supplier_key = normalize_key(supplier_name)
        if not system_key:
            return None, "Missing system_name."
        matches = []
        for row in self.master.system_by_name.get(system_key, []):
            supplier_id = str(row.get("Supplier ID", ""))
            supplier_rows = [s for s in self.master.suppliers if str(s.get("Supplier ID", "")) == supplier_id]
            supplier_name_value = supplier_rows[0]["Supplier Name"] if supplier_rows else ""
            if normalize_key(str(supplier_name_value)) == supplier_key:
                matches.append(row)
        if len(matches) == 1:
            return str(matches[0]["System ID"]), None
        if len(matches) > 1:
            return None, f"Ambiguous system match for {system_name}."
        return None, None

    def find_existing_component_id(self, component_name: str, system_id: str) -> tuple[str | None, str | None]:
        component_key = normalize_key(component_name)
        if not component_key:
            return None, "Missing component_name."
        matches = [
            row for row in self.master.component_by_name.get(component_key, [])
            if str(row.get("System ID", "")) == system_id
        ]
        if len(matches) == 1:
            return str(matches[0]["Component ID"]), None
        if len(matches) > 1:
            return None, f"Ambiguous component match for {component_name}."
        return None, None

    def ensure_supplier(self, supplier_name: str) -> tuple[str | None, str | None]:
        key = normalize_key(supplier_name)
        if not key:
            return None, "Missing supplier_name."
        existing = self.master.supplier_by_name.get(key, [])
        if len(existing) == 1:
            return str(existing[0]["Supplier ID"]), None
        if len(existing) > 1:
            return None, f"Ambiguous supplier match for {supplier_name}."
        if key in self.generated_suppliers:
            return self.generated_suppliers[key][0], None

        supplier_id = make_id("SUP", supplier_name, self.used_supplier_ids)
        row = {
            "Supplier ID": supplier_id,
            "Supplier Name": normalize_text(supplier_name),
            "Supplier Country": "",
            "Supplier Notes": ""
        }
        self.ready["SUPPLIERS"].append(row)
        self.generated_suppliers[key] = [supplier_id, normalize_text(supplier_name)]
        return supplier_id, None

    def ensure_system(self, system_name: str, supplier_name: str, category: str = "", notes: str = "") -> tuple[str | None, str | None]:
        system_key = normalize_key(system_name)
        supplier_key = normalize_key(supplier_name)
        if not system_key:
            return None, "Missing system_name."
        existing_id, existing_error = self.find_existing_system_id(system_name, supplier_name)
        if existing_id or existing_error:
            return existing_id, existing_error
        generated_key = f"{system_key}|{supplier_key}"
        if generated_key in self.generated_systems:
            return self.generated_systems[generated_key][0], None

        supplier_id, supplier_error = self.ensure_supplier(supplier_name)
        if supplier_error or not supplier_id:
            return None, supplier_error or "Unable to resolve supplier."

        system_id = make_id("SYS", f"{supplier_name}_{system_name}", self.used_system_ids)
        row = {
            "System ID": system_id,
            "System Name": normalize_text(system_name),
            "Supplier ID": supplier_id,
            "System Category": normalize_text(category),
            "System Notes": normalize_text(notes),
        }
        self.ready["SYSTEMS"].append(row)
        self.generated_systems[generated_key] = [system_id, normalize_text(system_name)]
        return system_id, None

    def ensure_component(self, component_name: str, system_name: str, component_role: str = "", implant_category: str = "", notes: str = "") -> tuple[str | None, str | None]:
        component_key = normalize_key(component_name)
        if not component_key:
            return None, "Missing component_name."
        system_id, system_error = self.ensure_system(system_name, self.payload.get("source_meta", {}).get("supplier_name", ""))
        if system_error or not system_id:
            return None, system_error or "Unable to resolve system."
        existing_id, existing_error = self.find_existing_component_id(component_name, system_id)
        if existing_id or existing_error:
            return existing_id, existing_error

        generated_key = f"{component_key}|{system_id}"
        if generated_key in self.generated_components:
            return self.generated_components[generated_key][0], None

        component_id = make_id("COMP", f"{system_name}_{component_name}", self.used_component_ids)
        row = {
            "Component ID": component_id,
            "Component Name": normalize_text(component_name),
            "Component Role": normalize_text(component_role),
            "Implant Category": normalize_text(implant_category),
            "System ID": system_id,
            "Component Notes": normalize_text(notes),
        }
        self.ready["COMPONENTS"].append(row)
        self.generated_components[generated_key] = [component_id, normalize_text(component_name)]
        return component_id, None

    def resolve_procedure_id(self, procedure_name: str, procedure_variant: str) -> tuple[str | None, str | None]:
        name_key = normalize_key(procedure_name)
        variant_key = normalize_key(procedure_variant)
        if not name_key:
            return None, "Missing procedure_name."
        exact_matches = self.master.procedure_by_name_variant.get((name_key, variant_key), [])
        if len(exact_matches) == 1:
            return str(exact_matches[0]["Procedure ID"]), None
        if len(exact_matches) > 1:
            return None, f"Ambiguous procedure match for {procedure_name} / {procedure_variant}."
        if not variant_key:
            broad_matches = self.master.procedure_by_name_only.get(name_key, [])
            procedure_ids = sorted({str(row["Procedure ID"]) for row in broad_matches if row.get("Procedure ID")})
            if len(procedure_ids) == 1:
                return procedure_ids[0], None
            if len(procedure_ids) > 1:
                return None, f"Procedure variant required to resolve {procedure_name}."
        return None, f"Procedure not found for {procedure_name} / {procedure_variant}."

    def process_systems(self) -> None:
        for row in self.payload.get("systems", []):
            normalized = {
                "system_name": normalize_text(row.get("system_name", "")),
                "supplier_name": normalize_text(row.get("supplier_name", "")),
                "system_category": normalize_text(row.get("system_category", "")),
                "system_notes": normalize_text(row.get("system_notes", "")),
            }
            missing = [field for field in ("system_name", "supplier_name") if not normalized[field]]
            if missing:
                self.missing_field_warnings.append({"entity": "SYSTEMS", "row": normalized, "missing_fields": missing})
                self.add_issue(self.rejected, "SYSTEMS", normalized, f"Missing required fields: {', '.join(missing)}")
                continue
            if any(
                normalize_key(str(existing.get("System Name", ""))) == normalize_key(normalized["system_name"])
                and normalize_key(str(sup.get("Supplier Name", ""))) == normalize_key(normalized["supplier_name"])
                for existing in self.master.systems
                for sup in self.master.suppliers
                if str(sup.get("Supplier ID", "")) == str(existing.get("Supplier ID", ""))
            ):
                self.duplicate_warnings.append({"entity": "SYSTEMS", "row": normalized, "reason": "System already exists in master dataset."})
                self.add_issue(self.needs_review, "SYSTEMS", normalized, "Duplicate system already exists in master dataset.")
                continue
            system_id, error = self.ensure_system(
                normalized["system_name"],
                normalized["supplier_name"],
                normalized["system_category"],
                normalized["system_notes"],
            )
            if error or not system_id:
                self.bad_link_warnings.append({"entity": "SYSTEMS", "row": normalized, "reason": error or "Unable to create system."})
                self.add_issue(self.rejected, "SYSTEMS", normalized, error or "Unable to create system.")

    def process_trays(self) -> None:
        for row in self.payload.get("trays", []):
            normalized = {
                "tray_name": normalize_text(row.get("tray_name", "")),
                "system_name": normalize_text(row.get("system_name", "")),
                "tray_type": normalize_text(row.get("tray_type", "")),
                "tray_notes": normalize_text(row.get("tray_notes", "")),
            }
            missing = [field for field in ("tray_name", "system_name") if not normalized[field]]
            if missing:
                self.missing_field_warnings.append({"entity": "TRAYS", "row": normalized, "missing_fields": missing})
                self.add_issue(self.rejected, "TRAYS", normalized, f"Missing required fields: {', '.join(missing)}")
                continue
            system_id, error = self.ensure_system(
                normalized["system_name"],
                self.payload.get("source_meta", {}).get("supplier_name", ""),
            )
            if error or not system_id:
                self.bad_link_warnings.append({"entity": "TRAYS", "row": normalized, "reason": error or "Unable to resolve system link."})
                self.add_issue(self.rejected, "TRAYS", normalized, error or "Unable to resolve system link.")
                continue
            duplicate = any(
                normalize_key(str(existing.get("Tray Name", ""))) == normalize_key(normalized["tray_name"])
                and str(existing.get("System ID", "")) == system_id
                for existing in self.master.trays
            )
            if duplicate:
                self.duplicate_warnings.append({"entity": "TRAYS", "row": normalized, "reason": "Tray already exists in master dataset."})
                self.add_issue(self.needs_review, "TRAYS", normalized, "Duplicate tray already exists in master dataset.")
                continue
            tray_id = make_id("TRAY", f"{normalized['system_name']}_{normalized['tray_name']}", self.used_tray_ids)
            self.ready["TRAYS"].append(
                {
                    "Tray ID": tray_id,
                    "Tray Name": normalized["tray_name"],
                    "System ID": system_id,
                    "Tray Type": normalized["tray_type"],
                    "Tray Notes": normalized["tray_notes"],
                }
            )

    def process_components(self) -> None:
        for row in self.payload.get("components", []):
            normalized = {
                "component_name": normalize_text(row.get("component_name", "")),
                "system_name": normalize_text(row.get("system_name", "")),
                "component_role": normalize_text(row.get("component_role", "")),
                "implant_category": normalize_text(row.get("implant_category", "")),
                "component_notes": normalize_text(row.get("component_notes", "")),
            }
            missing = [field for field in ("component_name", "system_name") if not normalized[field]]
            if missing:
                self.missing_field_warnings.append({"entity": "COMPONENTS", "row": normalized, "missing_fields": missing})
                self.add_issue(self.rejected, "COMPONENTS", normalized, f"Missing required fields: {', '.join(missing)}")
                continue
            component_id, error = self.ensure_component(
                normalized["component_name"],
                normalized["system_name"],
                normalized["component_role"],
                normalized["implant_category"],
                normalized["component_notes"],
            )
            if error or not component_id:
                self.bad_link_warnings.append({"entity": "COMPONENTS", "row": normalized, "reason": error or "Unable to resolve component link."})
                self.add_issue(self.rejected, "COMPONENTS", normalized, error or "Unable to resolve component link.")
                continue
            system_id, system_error = self.ensure_system(
                normalized["system_name"],
                self.payload.get("source_meta", {}).get("supplier_name", ""),
            )
            if system_error or not system_id:
                self.bad_link_warnings.append({"entity": "COMPONENTS", "row": normalized, "reason": system_error or "Unable to resolve system link."})
                self.add_issue(self.rejected, "COMPONENTS", normalized, system_error or "Unable to resolve system link.")
                continue
            role_duplicate = (
                normalized["component_role"]
                and any(
                    str(existing.get("System ID", "")) == system_id
                    and normalize_key(str(existing.get("Component Role", ""))) == normalize_key(normalized["component_role"])
                    for existing in self.master.components
                )
            )
            if role_duplicate:
                self.duplicate_warnings.append(
                    {
                        "entity": "COMPONENTS",
                        "row": normalized,
                        "reason": "Possible component role duplicate for this system.",
                    }
                )
                self.add_issue(
                    self.needs_review,
                    "COMPONENTS",
                    normalized,
                    "Possible component role duplicate for this system.",
                )
                continue
            existing_component_id, existing_error = self.find_existing_component_id(
                normalized["component_name"],
                system_id,
            )
            if existing_component_id and existing_component_id == component_id:
                self.duplicate_warnings.append({"entity": "COMPONENTS", "row": normalized, "reason": "Component already exists in master dataset."})
                self.add_issue(self.needs_review, "COMPONENTS", normalized, "Duplicate component already exists in master dataset.")

    def process_products(self) -> None:
        for row in self.payload.get("products", []):
            normalized = {
                "sku": normalize_text(row.get("sku", "")),
                "product_name": normalize_text(row.get("product_name", "")),
                "component_name": normalize_text(row.get("component_name", "")),
                "packaging_type": normalize_text(row.get("packaging_type", "")),
                "product_notes": normalize_text(row.get("product_notes", "")),
            }
            missing = []
            if not normalized["sku"] and not normalized["product_name"]:
                missing.append("sku or product_name")
            if not normalized["component_name"]:
                missing.append("component_name")
            if missing:
                self.missing_field_warnings.append({"entity": "PRODUCTS", "row": normalized, "missing_fields": missing})
                self.add_issue(self.rejected, "PRODUCTS", normalized, f"Missing required fields: {', '.join(missing)}")
                continue
            component_id, error = self.ensure_component(
                normalized["component_name"],
                self.payload.get("source_meta", {}).get("system_name", ""),
            )
            if error or not component_id:
                self.bad_link_warnings.append({"entity": "PRODUCTS", "row": normalized, "reason": error or "Unable to resolve component link."})
                self.add_issue(self.rejected, "PRODUCTS", normalized, error or "Unable to resolve component link.")
                continue
            duplicate = any(
                (normalized["sku"] and normalize_key(str(existing.get("SKU", ""))) == normalize_key(normalized["sku"]))
                or (
                    normalize_key(str(existing.get("Product Name", ""))) == normalize_key(normalized["product_name"])
                    and str(existing.get("Component ID", "")) == component_id
                )
                for existing in self.master.products
            )
            if duplicate:
                self.duplicate_warnings.append({"entity": "PRODUCTS", "row": normalized, "reason": "Product already exists in master dataset."})
                self.add_issue(self.needs_review, "PRODUCTS", normalized, "Duplicate product already exists in master dataset.")
                continue
            self.ready["PRODUCTS"].append(
                {
                    "SKU": normalized["sku"],
                    "Product Name": normalized["product_name"],
                    "Component ID": component_id,
                    "Packaging Type": normalized["packaging_type"],
                    "Product Notes": normalized["product_notes"],
                }
            )

    def process_system_mappings(self) -> None:
        for row in self.payload.get("system_mappings", []):
            normalized = {
                "procedure_name": normalize_text(row.get("procedure_name", "")),
                "procedure_variant": normalize_text(row.get("procedure_variant", "")),
                "system_name": normalize_text(row.get("system_name", "")),
                "data_source": normalize_text(row.get("data_source", "")) or normalize_text(self.payload.get("source_meta", {}).get("source_document", "")),
                "research_notes": normalize_text(row.get("research_notes", "")),
                "last_verified": normalize_text(row.get("last_verified", "")),
            }
            missing = [field for field in ("procedure_name", "system_name") if not normalized[field]]
            if missing:
                self.missing_field_warnings.append({"entity": "SYSTEM_MAPPINGS", "row": normalized, "missing_fields": missing})
                self.add_issue(self.rejected, "SYSTEM_MAPPINGS", normalized, f"Missing required fields: {', '.join(missing)}")
                continue
            procedure_id, procedure_error = self.resolve_procedure_id(
                normalized["procedure_name"],
                normalized["procedure_variant"],
            )
            system_id, system_error = self.ensure_system(
                normalized["system_name"],
                self.payload.get("source_meta", {}).get("supplier_name", ""),
            )
            if procedure_error or not procedure_id:
                self.bad_link_warnings.append({"entity": "SYSTEM_MAPPINGS", "row": normalized, "reason": procedure_error or "Unable to resolve procedure link."})
                self.add_issue(self.rejected, "SYSTEM_MAPPINGS", normalized, procedure_error or "Unable to resolve procedure link.")
                continue
            if system_error or not system_id:
                self.bad_link_warnings.append({"entity": "SYSTEM_MAPPINGS", "row": normalized, "reason": system_error or "Unable to resolve system link."})
                self.add_issue(self.rejected, "SYSTEM_MAPPINGS", normalized, system_error or "Unable to resolve system link.")
                continue
            duplicate = any(
                str(existing.get("Procedure ID", "")) == procedure_id
                and str(existing.get("System ID", "")) == system_id
                for existing in self.master.mappings
            )
            if duplicate:
                self.duplicate_warnings.append({"entity": "SYSTEM_MAPPINGS", "row": normalized, "reason": "System mapping already exists in master dataset."})
                self.add_issue(self.needs_review, "SYSTEM_MAPPINGS", normalized, "Duplicate system mapping already exists in master dataset.")
                continue
            mapping_id = make_id("MAP", f"{procedure_id}_{system_id}", self.used_mapping_ids)
            self.ready["SYSTEM_MAPPINGS"].append(
                {
                    "Mapping ID": mapping_id,
                    "Procedure ID": procedure_id,
                    "System ID": system_id,
                    "Mapping Status": "",
                    "Data Source": normalized["data_source"],
                    "Research Notes": normalized["research_notes"],
                    "Last Verified": normalized["last_verified"],
                }
            )

    def run(self) -> dict[str, Any]:
        self.process_systems()
        self.process_trays()
        self.process_components()
        self.process_products()
        self.process_system_mappings()
        return {
            "source_meta": self.payload.get("source_meta", {}),
            "ready_to_import": {sheet: self.ready.get(sheet, []) for sheet in TARGET_SHEETS},
            "needs_review": {sheet: self.needs_review.get(sheet, []) for sheet in TARGET_SHEETS},
            "rejected": {sheet: self.rejected.get(sheet, []) for sheet in TARGET_SHEETS},
            "duplicate_warnings": self.duplicate_warnings,
            "missing_field_warnings": self.missing_field_warnings,
            "bad_link_warnings": self.bad_link_warnings,
            "exception_report": self.exception_report,
        }


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate PrepSight fixed-data extraction JSON against the master workbook.")
    parser.add_argument("input_json", type=Path, help="Path to extraction JSON.")
    parser.add_argument("--master", type=Path, default=DEFAULT_MASTER_WORKBOOK, help="Path to master workbook.")
    parser.add_argument("--schema", type=Path, default=DEFAULT_SCHEMA, help="Path to locked extraction schema.")
    parser.add_argument("--output", type=Path, default=None, help="Path to output JSON report.")
    args = parser.parse_args()

    payload = load_json(args.input_json)
    schema = load_json(args.schema)
    schema_errors = validate_schema_shape(payload, schema)
    if schema_errors:
        report = {
            "schema_errors": schema_errors,
            "ready_to_import": {},
            "needs_review": {},
            "rejected": {},
            "duplicate_warnings": [],
            "missing_field_warnings": [],
            "bad_link_warnings": [],
            "exception_report": [{"entity": "SCHEMA", "reason": error} for error in schema_errors],
        }
    else:
        master = load_master_indexes(args.master)
        importer = Importer(master, payload)
        report = importer.run()

    output_path = args.output or args.input_json.with_name(f"{args.input_json.stem}_validated.json")
    output_path.write_text(json.dumps(report, indent=2, ensure_ascii=True), encoding="utf-8")
    print(output_path)


if __name__ == "__main__":
    main()
