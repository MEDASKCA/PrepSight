from __future__ import annotations

import json
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SYSTEMS_PATH = ROOT / "data" / "systems" / "systems.json"
SUPPLIERS_PATH = ROOT / "data" / "suppliers" / "suppliers.json"
LIVE_MAPPING_PATH = ROOT / "data" / "systems" / "trauma_and_orthopaedics_full_live_mapping.json"
OUTPUT_PATH = ROOT / "PrepSight_Fixed_Data_Inventory.xlsx"


SHEETS: OrderedDict[str, list[str]] = OrderedDict(
    {
        "PROCEDURES": [
            "Procedure ID",
            "Specialty",
            "Subspecialty",
            "Anatomy",
            "Subanatomy",
            "Procedure",
            "Procedure Variant",
        ],
        "SUPPLIERS": [
            "Supplier ID",
            "Supplier Name",
            "Supplier Country",
            "Supplier Notes",
        ],
        "SYSTEMS": [
            "System ID",
            "System Name",
            "Supplier ID",
            "System Category",
            "System Notes",
        ],
        "TRAYS": [
            "Tray ID",
            "Tray Name",
            "System ID",
            "Tray Type",
            "Tray Notes",
        ],
        "COMPONENTS": [
            "Component ID",
            "Component Name",
            "Component Role",
            "Implant Category",
            "System ID",
            "Component Notes",
        ],
        "PRODUCTS": [
            "SKU",
            "Product Name",
            "Component ID",
            "Packaging Type",
            "Product Notes",
        ],
        "SYSTEM_MAPPINGS": [
            "Mapping ID",
            "Procedure ID",
            "System ID",
            "Mapping Status",
            "Data Source",
            "Research Notes",
            "Last Verified",
        ],
    }
)


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def autosize_columns(worksheet) -> None:
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 44)


def style_sheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    if worksheet.max_column and worksheet.max_row:
        worksheet.auto_filter.ref = worksheet.dimensions
    autosize_columns(worksheet)


def build_procedure_rows(live_mapping: list[dict]) -> list[list[str]]:
    rows: list[list[str]] = []
    seen: set[tuple[str, str]] = set()
    for branch in live_mapping:
        specialty = branch.get("specialty", "")
        subspecialty = branch.get("subspecialty", "")
        anatomy = branch.get("anatomy", "")
        for subanatomy_group in branch.get("subanatomy_groups", []):
            subanatomy = subanatomy_group.get("name", "")
            for procedure in subanatomy_group.get("procedures", []):
                procedure_id = procedure.get("procedure_id", "")
                procedure_name = procedure.get("procedure_name", "")
                variants = procedure.get("variants", [])
                if variants:
                    for variant in variants:
                        variant_name = variant.get("variant_name", "")
                        key = (procedure_id, variant_name)
                        if key in seen:
                            continue
                        seen.add(key)
                        rows.append(
                            [
                                procedure_id,
                                specialty,
                                subspecialty,
                                anatomy,
                                subanatomy,
                                procedure_name,
                                variant_name,
                            ]
                        )
                else:
                    key = (procedure_id, "")
                    if key in seen:
                        continue
                    seen.add(key)
                    rows.append(
                        [
                            procedure_id,
                            specialty,
                            subspecialty,
                            anatomy,
                            subanatomy,
                            procedure_name,
                            "",
                        ]
                    )
    return rows


def build_supplier_rows(suppliers: list[dict]) -> list[list[str]]:
    rows: list[list[str]] = []
    for supplier in suppliers:
        rows.append(
            [
                supplier.get("id", ""),
                supplier.get("name", ""),
                "",
                "",
            ]
        )
    return rows


def build_system_rows(systems: list[dict]) -> list[list[str]]:
    rows: list[list[str]] = []
    for system in systems:
        rows.append(
            [
                system.get("id", ""),
                system.get("name", ""),
                system.get("supplier_id", ""),
                system.get("category", ""),
                system.get("description", ""),
            ]
        )
    return rows


def build_mapping_rows(live_mapping: list[dict]) -> list[list[str]]:
    rows: list[list[str]] = []
    for branch in live_mapping:
        data_source = branch.get("source_seed", "")
        for subanatomy_group in branch.get("subanatomy_groups", []):
            for procedure in subanatomy_group.get("procedures", []):
                procedure_id = procedure.get("procedure_id", "")
                for variant in procedure.get("variants", []):
                    variant_id = variant.get("variant_id", "")
                    for system in variant.get("systems", []):
                        system_id = system.get("system_id", "")
                        mapping_id = f"MAP_{variant_id}_{system_id}" if variant_id and system_id else ""
                        rows.append(
                            [
                                mapping_id,
                                procedure_id,
                                system_id,
                                system.get("mapping_status", ""),
                                data_source,
                                system.get("notes", ""),
                                "",
                            ]
                        )
    return rows


def write_sheet(worksheet, headers: list[str], rows: list[list[str]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    style_sheet(worksheet)


def main() -> None:
    suppliers = load_json(SUPPLIERS_PATH)
    systems = load_json(SYSTEMS_PATH)
    live_mapping = load_json(LIVE_MAPPING_PATH)

    workbook = Workbook()
    first_sheet = True
    for title, headers in SHEETS.items():
        if first_sheet:
            worksheet = workbook.active
            worksheet.title = title
            first_sheet = False
        else:
            worksheet = workbook.create_sheet(title=title)

        if title == "PROCEDURES":
            rows = build_procedure_rows(live_mapping)
        elif title == "SUPPLIERS":
            rows = build_supplier_rows(suppliers)
        elif title == "SYSTEMS":
            rows = build_system_rows(systems)
        elif title == "SYSTEM_MAPPINGS":
            rows = build_mapping_rows(live_mapping)
        else:
            rows = []

        write_sheet(worksheet, headers, rows)

    workbook.properties.creator = "Codex"
    workbook.properties.modified = datetime.now(timezone.utc).replace(tzinfo=None)
    workbook.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
