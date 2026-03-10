"""
Sync surgical_master_ontology.xlsx into the app JSON data layer.

This importer is additive:
- it merges missing records into taxonomy/procedure/variant/system/supplier JSON
- it does not overwrite existing records
- it normalises known supplier/specialty ID drift from the workbook

Run:
  python scripts/sync_ontology_from_excel.py
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
XLSX = Path(r"C:\Users\forda\OneDrive\Desktop\surgical_master_ontology.xlsx")
DATA = ROOT / "data"

SUPPLIER_NORMALIZATION = {
    "SUP_DEPUY": "SUP_DEPUY_SYNTHES",
    "SUP_ZIMMER": "SUP_ZIMMER_BIOMET",
    "SUP_SMITHNEPHEW": "SUP_SMITH_NEPHEW",
    "SUP_GLOBUS": "SUP_GLOBUS_MEDICAL",
    "SUP_BOSTON": "SUP_BOSTON_SCIENTIFIC",
    "SUP_PARAGON_28": "SUP_PARAGON28",
    "SUP_ORTHOPEDS": "SUP_ORTHOPEDIATRICS",
    "SUP_LED_A": "SUP_LEDA",
    "SUP_BBRAUN_AESCULAP": "SUP_BBRAUN",
    "SUP_B_BRAUN": "SUP_BBRAUN",
    "SUP_DRAEGER": "SUP_DRAGER",
    "SUP_EDWARDS": "SUP_EDWARDS_LIFESCIENCES",
    "SUP_INTEGRA": "SUP_INTEGRA_LIFESCIENCES",
    "SUP_LUMENIS": "SUP_MEDTRONIC",
    "SUP_FUJIFILM_SONOSITE": "SUP_FUJI_FILM_MEDICAL",
    "SUP_TORNIER": "SUP_SMITH_NEPHEW",
    "SUP_STANMORE_STRYKER": "SUP_STRYKER",
    "SUP_LEMaitre": "SUP_LEMaitre_VASCULAR",
}

SPECIALTY_NORMALIZATION = {
    "SPEC_GYNAECOLOGY": "SPEC_GYNECOLOGY",
    "SPEC_OTOLARYNGOLOGY": "SPEC_ENT",
    "SPEC_ORAL_AND_MAXILLOFACIAL": "SPEC_OMFS",
    "SPEC_PLASTIC_RECONSTRUCTIVE": "SPEC_PLASTIC_RECONSTRUCTIVE_SURGERY",
    "SPEC_PLASTIC_AND_RECONSTRUCTIVE": "SPEC_PLASTIC_RECONSTRUCTIVE_SURGERY",
}

SPECIALTY_TO_FOLDER = {
    "SPEC_ANAESTHESIA": "anaesthesia",
    "SPEC_CARDIOTHORACIC_SURGERY": "cardiothoracic",
    "SPEC_DENTAL_AND_ORAL": "dental_and_oral",
    "SPEC_DENTAL_ORAL_SURGERY": "dental_and_oral",
    "SPEC_GENERAL_SURGERY": "general_surgery",
    "SPEC_GYNECOLOGY": "gynaecology",
    "SPEC_NEUROSURGERY": "neurosurgery",
    "SPEC_OBSTETRICS": "obstetrics",
    "SPEC_OPHTHALMOLOGY": "opthalmology",
    "SPEC_OMFS": "oral_and_maxillofacial",
    "SPEC_ENT": "otolaryngology",
    "SPEC_PAEDIATRIC_SURGERY": "paediatric",
    "SPEC_PLASTIC_RECONSTRUCTIVE_SURGERY": "plastic_and_reconstructive",
    "SPEC_PODIATRIC_SURGERY": "podiatric",
    "SPEC_TRAUMA_ORTHOPAEDICS": "trauma_and_orthopaedics",
    "SPEC_UROLOGY": "urology",
    "SPEC_VASCULAR_SURGERY": "vascular",
}


def load_json(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def merge_by_id(existing: list[dict[str, Any]], incoming: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    index = {row["id"]: row for row in existing if row.get("id")}
    added = 0
    for row in incoming:
      row_id = row.get("id")
      if not row_id or row_id in index:
          continue
      index[row_id] = row
      added += 1
    return list(index.values()), added


def append_unique(
    existing: list[dict[str, Any]],
    incoming: list[dict[str, Any]],
    key_fn,
) -> tuple[list[dict[str, Any]], int]:
    seen = {key_fn(row) for row in existing}
    added = 0
    for row in incoming:
        key = key_fn(row)
        if key in seen:
            continue
        existing.append(row)
        seen.add(key)
        added += 1
    return existing, added


def normalize_supplier_id(value: Any) -> str | None:
    if not value:
        return None
    text = str(value).strip()
    if text.startswith("SUPP_"):
        return None
    return SUPPLIER_NORMALIZATION.get(text, text)


def normalize_specialty_id(value: Any) -> str:
    text = str(value or "").strip()
    return SPECIALTY_NORMALIZATION.get(text, text)


def split_aliases(value: Any) -> list[str]:
    if not value:
        return []
    text = str(value).strip()
    if not text:
        return []
    raw_parts = text.replace(";", "|").replace(",", "|").split("|")
    return [part.strip() for part in raw_parts if part and part.strip()]


def to_int(value: Any, default: int = 10) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def read_workbook_rows() -> tuple[dict[str, int], list[tuple[Any, ...]]]:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["MASTER_ALL_SPECIALTIES"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {name: idx for idx, name in enumerate(header_row)}
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return headers, rows


def get(row: tuple[Any, ...], headers: dict[str, int], column: str) -> Any:
    idx = headers.get(column)
    if idx is None:
        return None
    return row[idx]


def main() -> None:
    headers, rows = read_workbook_rows()
    print(f"Read {len(rows)} workbook rows from {XLSX}")

    service_lines: dict[str, dict[str, Any]] = {}
    anatomy: dict[str, dict[str, Any]] = {}
    procedures_by_spec: dict[str, dict[str, dict[str, Any]]] = {}
    variants_by_spec: dict[str, dict[str, dict[str, Any]]] = {}
    suppliers: dict[str, dict[str, Any]] = {}
    systems: dict[str, dict[str, Any]] = {}
    mappings: dict[tuple[str, str], dict[str, Any]] = {}
    service_line_anatomy_map: dict[tuple[str, str], dict[str, Any]] = {}

    skipped_placeholder_variants = 0

    for row in rows:
        specialty_id = normalize_specialty_id(get(row, headers, "specialty_id"))
        if not specialty_id:
            continue

        service_line_id = get(row, headers, "service_line_id")
        service_line_name = get(row, headers, "service_line_name")
        anatomy_id = get(row, headers, "anatomy_id")
        anatomy_name = get(row, headers, "anatomy_name")
        procedure_id = get(row, headers, "procedure_id")
        procedure_name = get(row, headers, "procedure_name")
        procedure_aliases = split_aliases(get(row, headers, "procedure_aliases"))
        procedure_description = get(row, headers, "procedure_description") or ""
        procedure_status = get(row, headers, "procedure_status") or "active"

        if service_line_id and service_line_id not in service_lines:
            service_lines[service_line_id] = {
                "id": service_line_id,
                "name": service_line_name or service_line_id,
                "specialty_id": specialty_id,
            }

        if anatomy_id and anatomy_id not in anatomy:
            anatomy[anatomy_id] = {
                "id": anatomy_id,
                "name": anatomy_name or anatomy_id,
                "specialty_id": specialty_id,
                "parent_id": None,
                "sort_order": 10,
                "tags": [str(anatomy_name).lower()] if anatomy_name else [],
            }

        if service_line_id and anatomy_id and (service_line_id, anatomy_id) not in service_line_anatomy_map:
            service_line_anatomy_map[(service_line_id, anatomy_id)] = {
                "id": f"SLAMAP_{service_line_id}_{anatomy_id}",
                "service_line_id": service_line_id,
                "anatomy_id": anatomy_id,
                "sort_order": 10,
            }

        if procedure_id:
            procedures_by_spec.setdefault(specialty_id, {})
            if procedure_id not in procedures_by_spec[specialty_id]:
                procedures_by_spec[specialty_id][procedure_id] = {
                    "id": procedure_id,
                    "name": procedure_name or procedure_id,
                    "specialty_id": specialty_id,
                    "service_line_id": service_line_id or "",
                    "anatomy_id": anatomy_id or "",
                    "aliases": procedure_aliases,
                    "description": procedure_description,
                    "status": procedure_status if procedure_status in {"active", "inactive"} else "active",
                }

        supplier_id = normalize_supplier_id(get(row, headers, "supplier_id"))
        supplier_name = get(row, headers, "supplier_name") or ""
        supplier_aliases = split_aliases(get(row, headers, "supplier_aliases"))
        if supplier_id and supplier_id not in suppliers:
            suppliers[supplier_id] = {
                "id": supplier_id,
                "name": supplier_name or supplier_id,
                "aliases": supplier_aliases,
                "status": "active",
            }

        system_id = get(row, headers, "system_id")
        system_name = get(row, headers, "system_name")
        system_aliases = split_aliases(get(row, headers, "system_aliases"))
        system_type = get(row, headers, "system_type") or ""
        system_category = get(row, headers, "system_category") or ""
        system_description = get(row, headers, "system_description") or ""
        system_status = get(row, headers, "system_status") or "active"
        if system_id and system_id not in systems:
            systems[system_id] = {
                "id": system_id,
                "name": system_name or system_id,
                "supplier_id": supplier_id,
                "specialty_id": specialty_id,
                "service_line_ids": [service_line_id] if service_line_id else [],
                "anatomy_ids": [anatomy_id] if anatomy_id else [],
                "system_type": system_type,
                "category": system_category,
                "aliases": system_aliases,
                "description": system_description,
                "status": system_status if system_status in {"active", "inactive"} else "active",
            }

        variant_id = get(row, headers, "variant_id")
        variant_name = get(row, headers, "variant_name")
        variant_type = get(row, headers, "variant_type") or "technique"
        data_quality_status = get(row, headers, "data_quality_status") or ""
        if variant_type == "technique_profile" and data_quality_status == "completed_with_inference":
            skipped_placeholder_variants += 1
            continue

        if variant_id and procedure_id:
            variants_by_spec.setdefault(specialty_id, {})
            variants_by_spec[specialty_id][variant_id] = {
                "id": variant_id,
                "procedure_id": procedure_id,
                "setting": get(row, headers, "setting") or "operating_theatre",
                "specialty_id": specialty_id,
                "service_line_id": service_line_id or "",
                "anatomy_id": anatomy_id or "",
                "name": variant_name or variant_id,
                "variant_type": variant_type,
                "variant_value": get(row, headers, "variant_value") or "",
                "description": get(row, headers, "variant_description") or "",
                "sort_order": to_int(get(row, headers, "variant_sort_order"), 10),
                "status": (get(row, headers, "variant_status") or "active"),
            }

        if variant_id and system_id:
            mapping_key = (variant_id, system_id)
            if mapping_key not in mappings:
                mappings[mapping_key] = {
                    "id": get(row, headers, "system_map_id") or f"MAP_{variant_id}_{system_id}",
                    "procedure_variant_id": variant_id,
                    "system_id": system_id,
                    "is_default": str(get(row, headers, "mapping_is_default") or "").lower() == "true",
                    "role": get(row, headers, "mapping_role") or "primary",
                    "usage_type": get(row, headers, "mapping_usage_type") or "required",
                    "status": (get(row, headers, "mapping_status") or "active"),
                }

    print(f"Skipped {skipped_placeholder_variants} placeholder/inferred variants")

    taxonomy_paths = {
        "service_lines": DATA / "taxonomy" / "service_lines.json",
        "anatomy": DATA / "taxonomy" / "anatomy.json",
        "service_line_anatomy_map": DATA / "taxonomy" / "service_line_anatomy_map.json",
    }

    merged, added = merge_by_id(
        load_json(taxonomy_paths["service_lines"]),
        list(service_lines.values()),
    )
    save_json(taxonomy_paths["service_lines"], merged)
    print(f"service_lines: +{added} -> {len(merged)} total")

    merged, added = merge_by_id(
        load_json(taxonomy_paths["anatomy"]),
        list(anatomy.values()),
    )
    save_json(taxonomy_paths["anatomy"], merged)
    print(f"anatomy: +{added} -> {len(merged)} total")

    merged_map, added = append_unique(
        load_json(taxonomy_paths["service_line_anatomy_map"]),
        list(service_line_anatomy_map.values()),
        lambda row: (row["service_line_id"], row["anatomy_id"]),
    )
    save_json(taxonomy_paths["service_line_anatomy_map"], merged_map)
    print(f"service_line_anatomy_map: +{added} -> {len(merged_map)} total")

    for specialty_id, procedure_map in sorted(procedures_by_spec.items()):
        folder = SPECIALTY_TO_FOLDER.get(specialty_id)
        if not folder:
            print(f"skip procedures for unknown specialty {specialty_id}")
            continue
        path = DATA / "procedures" / folder / f"procedures_{folder}.json"
        merged, added = merge_by_id(load_json(path), list(procedure_map.values()))
        save_json(path, merged)
        print(f"procedures/{folder}: +{added} -> {len(merged)} total")

    for specialty_id, variant_map in sorted(variants_by_spec.items()):
        folder = SPECIALTY_TO_FOLDER.get(specialty_id)
        if not folder:
            print(f"skip variants for unknown specialty {specialty_id}")
            continue
        path = DATA / "procedure_variants" / folder / f"procedure_variants_{folder}.json"
        merged, added = merge_by_id(load_json(path), list(variant_map.values()))
        save_json(path, merged)
        print(f"procedure_variants/{folder}: +{added} -> {len(merged)} total")

    path = DATA / "suppliers" / "suppliers.json"
    merged, added = merge_by_id(load_json(path), list(suppliers.values()))
    save_json(path, merged)
    print(f"suppliers: +{added} -> {len(merged)} total")

    path = DATA / "systems" / "systems.json"
    merged, added = merge_by_id(load_json(path), list(systems.values()))
    save_json(path, merged)
    print(f"systems: +{added} -> {len(merged)} total")

    path = DATA / "systems" / "procedure_variant_system_map.json"
    merged, added = merge_by_id(load_json(path), list(mappings.values()))
    save_json(path, merged)
    print(f"variant-system mappings: +{added} -> {len(merged)} total")


if __name__ == "__main__":
    main()
