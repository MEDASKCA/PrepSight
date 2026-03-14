from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MASTER_WORKBOOK = ROOT / "PrepSight_Fixed_Data_Inventory.xlsx"
TARGET_SHEETS = [
    "SUPPLIERS",
    "SYSTEMS",
    "TRAYS",
    "COMPONENTS",
    "PRODUCTS",
    "SYSTEM_MAPPINGS",
]
REPORT_SHEET = "IMPORT_REPORT"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def normalize_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def get_headers(worksheet) -> list[str]:
    return [normalize_value(cell.value) for cell in worksheet[1]]


def get_existing_rows(worksheet, headers: list[str]) -> set[tuple[str, ...]]:
    existing: set[tuple[str, ...]] = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        values = tuple(normalize_value(value) for value in row[: len(headers)])
        if any(values):
            existing.add(values)
    return existing


def autosize_columns(worksheet) -> None:
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = normalize_value(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 80)


def style_report_sheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_columns(worksheet)


def ensure_report_sheet(workbook):
    if REPORT_SHEET in workbook.sheetnames:
        del workbook[REPORT_SHEET]
    worksheet = workbook.create_sheet(REPORT_SHEET)
    headers = ["Category", "Sheet", "Status", "Reason", "Row Data"]
    worksheet.append(headers)
    return worksheet


def append_report_rows(worksheet, category: str, sheet_name: str, status: str, reason: str, rows: list[dict[str, Any]]) -> None:
    for row in rows:
        worksheet.append(
            [
                category,
                sheet_name,
                status,
                reason,
                json.dumps(row, ensure_ascii=True, sort_keys=True),
            ]
        )


def main() -> None:
    parser = argparse.ArgumentParser(description="Import ready-to-import fixed-data rows into a new workbook copy.")
    parser.add_argument("validated_json", type=Path, help="Path to validated extraction JSON.")
    parser.add_argument("--master", type=Path, default=DEFAULT_MASTER_WORKBOOK, help="Path to master workbook.")
    parser.add_argument("--output", type=Path, default=None, help="Path to output workbook copy.")
    args = parser.parse_args()

    validated = load_json(args.validated_json)
    workbook = load_workbook(args.master)
    report_sheet = ensure_report_sheet(workbook)

    imported_counts: dict[str, int] = {sheet: 0 for sheet in TARGET_SHEETS}
    skipped_counts: dict[str, int] = {sheet: 0 for sheet in TARGET_SHEETS}

    for sheet_name in TARGET_SHEETS:
        worksheet = workbook[sheet_name]
        headers = get_headers(worksheet)
        existing_rows = get_existing_rows(worksheet, headers)
        ready_rows = validated.get("ready_to_import", {}).get(sheet_name, []) or []

        for row in ready_rows:
            ordered_values = tuple(normalize_value(row.get(header, "")) for header in headers)
            if ordered_values in existing_rows:
                skipped_counts[sheet_name] += 1
                append_report_rows(report_sheet, "ready_to_import", sheet_name, "skipped_duplicate", "Row already exists in target sheet.", [row])
                continue
            worksheet.append(list(ordered_values))
            existing_rows.add(ordered_values)
            imported_counts[sheet_name] += 1
            append_report_rows(report_sheet, "ready_to_import", sheet_name, "imported", "Row appended to target sheet.", [row])

        needs_review_rows = validated.get("needs_review", {}).get(sheet_name, []) or []
        rejected_rows = validated.get("rejected", {}).get(sheet_name, []) or []
        append_report_rows(report_sheet, "needs_review", sheet_name, "not_imported", "Row requires manual review.", needs_review_rows)
        append_report_rows(report_sheet, "rejected", sheet_name, "not_imported", "Row rejected during validation.", rejected_rows)

    for warning_name, status, reason in [
        ("duplicate_warnings", "warning", "Duplicate warning from validator."),
        ("missing_field_warnings", "warning", "Missing required fields from validator."),
        ("bad_link_warnings", "warning", "Bad link warning from validator."),
        ("exception_report", "warning", "Validator exception report entry."),
    ]:
        for warning in validated.get(warning_name, []) or []:
            sheet_name = normalize_value(warning.get("entity", ""))
            report_sheet.append(
                [
                    warning_name,
                    sheet_name,
                    status,
                    reason,
                    json.dumps(warning, ensure_ascii=True, sort_keys=True),
                ]
            )

    style_report_sheet(report_sheet)

    output_path = args.output or args.master.with_name(f"{args.master.stem}_imported{args.master.suffix}")
    workbook.save(output_path)
    workbook.close()

    summary = {
        "output_workbook": str(output_path),
        "imported_rows": imported_counts,
        "skipped_duplicates": skipped_counts,
    }
    print(json.dumps(summary, indent=2, ensure_ascii=True))


if __name__ == "__main__":
    main()
