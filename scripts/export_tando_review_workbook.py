from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path("C:/Users/forda/orthopod")
LIVE_MAPPING_PATH = ROOT / "data/systems/trauma_and_orthopaedics_full_live_mapping.json"
UNCERTAIN_PATH = ROOT / "data/systems/trauma_and_orthopaedics_uncertain_system_placements.json"
OUTPUT_PATH = ROOT / "Trauma_and_Orthopaedics_Review.xlsx"
FALLBACK_OUTPUT_PATH = ROOT / "Trauma_and_Orthopaedics_Review_v2.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="0EA5E9")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CONFIRMED_FILL = PatternFill("solid", fgColor="DCFCE7")
REVIEW_FILL = PatternFill("solid", fgColor="FEF3C7")
REJECT_FILL = PatternFill("solid", fgColor="FEE2E2")
DISCUSS_FILL = PatternFill("solid", fgColor="EDE9FE")
LEGACY_FILL = PatternFill("solid", fgColor="E5E7EB")
ROW_ALERT_FILL = PatternFill("solid", fgColor="FEE2E2")
ROW_WARN_FILL = PatternFill("solid", fgColor="FEF3C7")


def read_json(path: Path):
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def auto_fit(ws):
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for index, value in enumerate(row, start=1):
            if value is None:
                continue
            widths[index] = min(max(widths.get(index, 0), len(str(value)) + 2), 60)
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width


def style_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.sheet_view.showGridLines = True


def add_list_sheet(wb: Workbook):
    ws = wb.create_sheet("Lists")
    ws.sheet_state = "hidden"

    option_sets = {
        "Mapping Status": ["confirmed", "review_required", "rejected"],
        "Review Decision": ["leave_as_is", "confirm", "remove", "needs_discussion"],
        "Fixation Class": [
            "cemented",
            "cementless",
            "hybrid",
            "mobile_bearing",
            "fixed_bearing",
            "stemmed",
            "stemless",
            "linked",
            "unlinked",
            "reverse",
            "hemi",
            "resurfacing",
            "tumour",
            "salvage",
            "custom",
            "unknown",
        ],
        "Active Status": ["active", "legacy", "unknown"],
        "Approve": ["", "yes", "no"],
        "Keep": ["", "yes", "no"],
        "Remove": ["", "yes", "no"],
    }

    for col_index, (title, values) in enumerate(option_sets.items(), start=1):
        ws.cell(row=1, column=col_index, value=title)
        for row_index, value in enumerate(values, start=2):
            ws.cell(row=row_index, column=col_index, value=value)

    return option_sets


def add_validation(ws, target_range: str, list_col: int, list_length: int):
    col = get_column_letter(list_col)
    formula = f"=Lists!${col}$2:${col}${list_length + 1}"
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.prompt = "Choose from the dropdown."
    validation.error = "Select a value from the list."
    ws.add_data_validation(validation)
    validation.add(target_range)


def add_status_formatting(ws, header_map: dict[str, int], max_row: int):
    mapping_col = get_column_letter(header_map["Mapping Status"])
    decision_col = get_column_letter(header_map["Review Decision"])
    active_col = get_column_letter(header_map["Active Status"])
    keep_col = get_column_letter(header_map.get("Keep", 0)) if "Keep" in header_map else None
    remove_col = get_column_letter(header_map.get("Remove", 0)) if "Remove" in header_map else None

    ws.conditional_formatting.add(
        f"{mapping_col}2:{mapping_col}{max_row}",
        FormulaRule(formula=[f'${mapping_col}2="confirmed"'], fill=CONFIRMED_FILL),
    )
    ws.conditional_formatting.add(
        f"{mapping_col}2:{mapping_col}{max_row}",
        FormulaRule(formula=[f'${mapping_col}2="review_required"'], fill=REVIEW_FILL),
    )
    ws.conditional_formatting.add(
        f"{mapping_col}2:{mapping_col}{max_row}",
        FormulaRule(formula=[f'${mapping_col}2="rejected"'], fill=REJECT_FILL),
    )
    ws.conditional_formatting.add(
        f"{decision_col}2:{decision_col}{max_row}",
        FormulaRule(formula=[f'${decision_col}2="confirm"'], fill=CONFIRMED_FILL),
    )
    ws.conditional_formatting.add(
        f"{decision_col}2:{decision_col}{max_row}",
        FormulaRule(formula=[f'${decision_col}2="remove"'], fill=REJECT_FILL),
    )
    ws.conditional_formatting.add(
        f"{decision_col}2:{decision_col}{max_row}",
        FormulaRule(formula=[f'${decision_col}2="needs_discussion"'], fill=DISCUSS_FILL),
    )
    ws.conditional_formatting.add(
        f"{active_col}2:{active_col}{max_row}",
        FormulaRule(formula=[f'${active_col}2="legacy"'], fill=LEGACY_FILL),
    )

    if keep_col:
        ws.conditional_formatting.add(
            f"{keep_col}2:{keep_col}{max_row}",
            FormulaRule(formula=[f'${keep_col}2="yes"'], fill=CONFIRMED_FILL),
        )
    if remove_col:
        ws.conditional_formatting.add(
            f"{remove_col}2:{remove_col}{max_row}",
            FormulaRule(formula=[f'${remove_col}2="yes"'], fill=REJECT_FILL),
        )


def build_rows():
    live_mapping = read_json(LIVE_MAPPING_PATH)
    uncertain = read_json(UNCERTAIN_PATH)

    clinical_rows = []
    variant_rows = []
    mapping_rows = []
    uncertain_rows = []
    systems_master = {}

    for branch in live_mapping:
        for subanatomy in branch["subanatomy_groups"]:
            for procedure in subanatomy["procedures"]:
                system_count = sum(len(variant["systems"]) for variant in procedure["variants"])
                owner_note = ""
                if len(procedure["variants"]) == 0 and system_count == 0:
                    owner_note = "Needs variants added or classify as not implant-led."
                elif len(procedure["variants"]) == 0:
                    owner_note = "Needs variants added."
                elif system_count == 0:
                    owner_note = "Awaiting supplier mapping or no systems needed."
                clinical_rows.append(
                    [
                        branch["specialty"],
                        branch["subspecialty"],
                        branch["anatomy"],
                        subanatomy["name"],
                        procedure["procedure_id"],
                        procedure["procedure_name"],
                        len(procedure["variants"]),
                        system_count,
                        owner_note,
                    ]
                )

                for variant in procedure["variants"]:
                    variant_rows.append(
                        [
                            procedure["procedure_id"],
                            procedure["procedure_name"],
                            variant["variant_id"],
                            variant["variant_name"],
                            "Approach / Variant",
                            len(variant["systems"]),
                            "",
                        ]
                    )

                    for system in variant["systems"]:
                        mapping_rows.append(
                            [
                                procedure["procedure_id"],
                                variant["variant_id"],
                                system["system_id"],
                                branch["subspecialty"],
                                branch["anatomy"],
                                subanatomy["name"],
                                procedure["procedure_name"],
                                variant["variant_name"],
                                system["system_name"],
                                system["supplier_name"],
                                system["mapping_status"],
                                system["fixation_class"].replace("_", " ").title(),
                                system["active_status"],
                                "leave_as_is",
                                "",
                                "",
                                system.get("notes", ""),
                            ]
                        )

                        systems_master[system["system_id"]] = [
                            system["system_id"],
                            system["system_name"],
                            system["supplier_name"],
                            system.get("supplier_tier", ""),
                            system.get("procedure_class", "").replace("_", " ").title(),
                            system["fixation_class"].replace("_", " ").title(),
                            system["active_status"],
                            "",
                        ]

    for row in uncertain:
        for suggestion in row["suggested_systems"]:
            uncertain_rows.append(
                [
                    row["procedure_id"],
                    row["variant_id"],
                    suggestion["system_id"],
                    row["subspecialty"],
                    row["anatomy"],
                    row["subanatomy_group"],
                    row["procedure_name"],
                    row["variant_name"],
                    suggestion["system_name"],
                    suggestion["supplier_name"],
                    "review_required",
                    suggestion["procedure_class"].replace("_", " ").title(),
                    "leave_as_is",
                    "",
                    "",
                    suggestion.get("notes", ""),
                ]
            )

            systems_master.setdefault(
                suggestion["system_id"],
                [
                    suggestion["system_id"],
                    suggestion["system_name"],
                    suggestion["supplier_name"],
                    "",
                    suggestion.get("procedure_class", "").replace("_", " ").title(),
                    "",
                    "active",
                    "",
                ],
            )

    clinical_rows.sort(key=lambda row: tuple(str(item) for item in row[:6]))
    variant_rows.sort(key=lambda row: tuple(str(item) for item in row[:4]))
    mapping_rows.sort(key=lambda row: tuple(str(item) for item in row[:10]))
    uncertain_rows.sort(key=lambda row: tuple(str(item) for item in row[:10]))

    return clinical_rows, variant_rows, mapping_rows, uncertain_rows, sorted(
        systems_master.values(),
        key=lambda row: tuple(str(item) for item in row[:3]),
    )


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Clinical Tree"

    option_sets = add_list_sheet(wb)
    clinical_rows, variant_rows, mapping_rows, uncertain_rows, systems_rows = build_rows()

    sheet_specs = [
        (
            ws,
            [
                "Specialty",
                "Subspecialty",
                "Anatomy",
                "Subanatomy Group",
                "Procedure ID",
                "Procedure",
                "Variant Count",
                "System Count",
                "Owner Notes",
            ],
            clinical_rows,
        ),
        (
            wb.create_sheet("Variants"),
            [
                "Procedure ID",
                "Procedure",
                "Variant ID",
                "Variant",
                "Variant Type",
                "System Count",
                "Owner Notes",
            ],
            variant_rows,
        ),
        (
            wb.create_sheet("Live System Mappings"),
            [
                "Procedure ID",
                "Variant ID",
                "System ID",
                "Subspecialty",
                "Anatomy",
                "Subanatomy Group",
                "Procedure",
                "Variant",
                "System",
                "Supplier",
                "Mapping Status",
                "Fixation Class",
                "Active Status",
                "Review Decision",
                "Keep",
                "Remove",
                "Review Notes",
            ],
            mapping_rows,
        ),
        (
            wb.create_sheet("Suggested Review"),
            [
                "Procedure ID",
                "Variant ID",
                "Suggested System ID",
                "Subspecialty",
                "Anatomy",
                "Subanatomy Group",
                "Procedure",
                "Variant",
                "Suggested System",
                "Supplier",
                "Mapping Status",
                "Procedure Class",
                "Review Decision",
                "Approve",
                "Reject",
                "Review Notes",
            ],
            uncertain_rows,
        ),
        (
            wb.create_sheet("System Master"),
            [
                "System ID",
                "System",
                "Supplier",
                "Supplier Tier",
                "Procedure Class",
                "Fixation Class",
                "Active Status",
                "Notes",
            ],
            systems_rows,
        ),
    ]

    for current_ws, headers, rows in sheet_specs:
        current_ws.append(headers)
        for row in rows:
            current_ws.append(row)
        style_sheet(current_ws)
        auto_fit(current_ws)

    for sheet_name in ["Live System Mappings", "Suggested Review", "System Master"]:
        current_ws = wb[sheet_name]
        header_map = {cell.value: index for index, cell in enumerate(current_ws[1], start=1)}
        max_row = current_ws.max_row

        if "Mapping Status" in header_map:
            add_validation(current_ws, f"{get_column_letter(header_map['Mapping Status'])}2:{get_column_letter(header_map['Mapping Status'])}{max_row}", 1, len(option_sets["Mapping Status"]))
        if "Review Decision" in header_map:
            add_validation(current_ws, f"{get_column_letter(header_map['Review Decision'])}2:{get_column_letter(header_map['Review Decision'])}{max_row}", 2, len(option_sets["Review Decision"]))
        if "Fixation Class" in header_map:
            add_validation(current_ws, f"{get_column_letter(header_map['Fixation Class'])}2:{get_column_letter(header_map['Fixation Class'])}{max_row}", 3, len(option_sets["Fixation Class"]))
        if "Active Status" in header_map:
            add_validation(current_ws, f"{get_column_letter(header_map['Active Status'])}2:{get_column_letter(header_map['Active Status'])}{max_row}", 4, len(option_sets["Active Status"]))
        if "Approve" in header_map:
            add_validation(current_ws, f"{get_column_letter(header_map['Approve'])}2:{get_column_letter(header_map['Approve'])}{max_row}", 5, len(option_sets["Approve"]))
        if "Reject" in header_map:
            add_validation(current_ws, f"{get_column_letter(header_map['Reject'])}2:{get_column_letter(header_map['Reject'])}{max_row}", 5, len(option_sets["Approve"]))
        if "Keep" in header_map:
            add_validation(current_ws, f"{get_column_letter(header_map['Keep'])}2:{get_column_letter(header_map['Keep'])}{max_row}", 6, len(option_sets["Keep"]))
        if "Remove" in header_map:
            add_validation(current_ws, f"{get_column_letter(header_map['Remove'])}2:{get_column_letter(header_map['Remove'])}{max_row}", 7, len(option_sets["Remove"]))

        if "Mapping Status" in header_map and "Review Decision" in header_map and "Active Status" in header_map:
            add_status_formatting(current_ws, header_map, max_row)

    clinical_ws = wb["Clinical Tree"]
    variant_count_col = get_column_letter(7)
    system_count_col = get_column_letter(8)
    owner_notes_col = get_column_letter(9)
    clinical_ws.conditional_formatting.add(
        f"{system_count_col}2:{system_count_col}{clinical_ws.max_row}",
        CellIsRule(operator="equal", formula=["0"], fill=REJECT_FILL),
    )
    clinical_ws.conditional_formatting.add(
        f"{variant_count_col}2:{variant_count_col}{clinical_ws.max_row}",
        CellIsRule(operator="equal", formula=["0"], fill=REVIEW_FILL),
    )
    clinical_ws.conditional_formatting.add(
        f"A2:{owner_notes_col}{clinical_ws.max_row}",
        FormulaRule(
            formula=[f"AND(${variant_count_col}2=0,${system_count_col}2=0)"],
            fill=ROW_ALERT_FILL,
        ),
    )
    clinical_ws.conditional_formatting.add(
        f"A2:{owner_notes_col}{clinical_ws.max_row}",
        FormulaRule(
            formula=[f"AND(${variant_count_col}2=0,${system_count_col}2>0)"],
            fill=ROW_WARN_FILL,
        ),
    )
    clinical_ws.conditional_formatting.add(
        f"A2:{owner_notes_col}{clinical_ws.max_row}",
        FormulaRule(
            formula=[f"AND(${variant_count_col}2>0,${system_count_col}2=0)"],
            fill=ROW_ALERT_FILL,
        ),
    )

    try:
        wb.save(OUTPUT_PATH)
        print(OUTPUT_PATH)
    except PermissionError:
        wb.save(FALLBACK_OUTPUT_PATH)
        print(FALLBACK_OUTPUT_PATH)


if __name__ == "__main__":
    build_workbook()
