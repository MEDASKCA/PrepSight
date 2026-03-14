"""
Build PrepSight_Theatre_Equipment_Inventory.xlsx from the sourced JSON extraction files.
Run from any directory: python build_theatre_equipment_xlsx.py
"""

import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path("C:/Users/forda/orthopod")
EXTRACTIONS = BASE / "data/catalogue/manual_extractions/theatre-equipment"
OUTPUT = BASE / "PrepSight_Theatre_Equipment_Inventory.xlsx"

EXTRACTION_FILES = [
    ("operating_tables_extraction.json", "Operating Tables"),
    ("positioning_extraction.json", "Positioning"),
    ("consumables_extraction.json", "Consumables"),
    ("theatre_equipment_extraction.json", "Theatre Equipment"),
]

COLUMNS = [
    ("supplier", "Supplier"),
    ("product_family", "Product Family"),
    ("system_name", "System Name"),
    ("category", "Category"),
    ("subtype", "Subtype"),
    ("source_document", "Source Document"),
    ("source_url", "Source URL"),
    ("notes", "Notes"),
    ("product_codes", "Product Codes"),
    ("sourced", "Sourced"),
    ("sourced_date", "Sourced Date"),
    ("_subcategory", "Subcategory"),
]

HEADER_FILL = PatternFill("solid", fgColor="003366")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBCAT_FILLS = {
    "Operating Tables":   PatternFill("solid", fgColor="DDEEFF"),
    "Positioning":        PatternFill("solid", fgColor="E8F4E8"),
    "Consumables":        PatternFill("solid", fgColor="FFF4E0"),
    "Theatre Equipment":  PatternFill("solid", fgColor="F4E8FF"),
}
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def cell_value(rec, key, subcategory):
    if key == "_subcategory":
        return subcategory
    v = rec.get(key, "")
    if isinstance(v, list):
        return "; ".join(str(x) for x in v) if v else ""
    if isinstance(v, bool):
        return "Yes" if v else "No"
    return str(v) if v is not None else ""


def build_xlsx():
    wb = Workbook()

    # --- Summary sheet ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = "PrepSight Theatre Equipment Inventory"
    ws_sum["A1"].font = Font(bold=True, size=14, color="003366")
    ws_sum["A2"] = "Sourced Date: 2026-03-14"
    ws_sum["A3"] = "Source: Supplier websites and official documentation only"
    ws_sum["A4"] = ""
    ws_sum["A5"] = "Subcategory"
    ws_sum["B5"] = "Record Count"
    ws_sum["A5"].font = Font(bold=True)
    ws_sum["B5"].font = Font(bold=True)
    row = 6
    all_records = []
    for fname, label in EXTRACTION_FILES:
        fpath = EXTRACTIONS / fname
        if not fpath.exists():
            print(f"  WARNING: {fpath} not found")
            continue
        with open(fpath, encoding="utf-8") as f:
            data = json.load(f)
        records = data.get("records", [])
        for r in records:
            r["_subcategory"] = label
        all_records.extend(records)
        ws_sum[f"A{row}"] = label
        ws_sum[f"B{row}"] = len(records)
        row += 1
    ws_sum[f"A{row}"] = "TOTAL"
    ws_sum[f"B{row}"] = len(all_records)
    ws_sum[f"A{row}"].font = Font(bold=True)
    ws_sum[f"B{row}"].font = Font(bold=True)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15

    # --- All Records sheet ---
    ws_all = wb.create_sheet("All Records")
    headers = [col[1] for col in COLUMNS]
    ws_all.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws_all.row_dimensions[1].height = 30

    for rec in all_records:
        subcategory = rec.get("_subcategory", "")
        row_vals = [cell_value(rec, col[0], subcategory) for col in COLUMNS]
        ws_all.append(row_vals)
        ri = ws_all.max_row
        fill = SUBCAT_FILLS.get(subcategory)
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws_all.cell(row=ri, column=col_idx)
            if fill:
                cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    widths = [22, 24, 34, 22, 30, 32, 50, 80, 30, 10, 14, 20]
    for i, w in enumerate(widths, 1):
        ws_all.column_dimensions[get_column_letter(i)].width = w

    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # --- Per-subcategory sheets ---
    for fname, label in EXTRACTION_FILES:
        fpath = EXTRACTIONS / fname
        if not fpath.exists():
            continue
        with open(fpath, encoding="utf-8") as f:
            data = json.load(f)
        records = data.get("records", [])
        sheet_name = label[:31]
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.row_dimensions[1].height = 30
        fill = SUBCAT_FILLS.get(label)
        for rec in records:
            row_vals = [cell_value(rec, col[0], label) for col in COLUMNS]
            ws.append(row_vals)
            ri = ws.max_row
            for col_idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=ri, column=col_idx)
                if fill:
                    cell.fill = fill
                cell.border = BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")
    print(f"Total records: {len(all_records)}")


if __name__ == "__main__":
    build_xlsx()
